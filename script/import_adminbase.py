"""Import ไฟล์ Excel ใน adminbase/ ไปยัง MongoDB collection knowledge_base.

หลักการสำคัญ:
- ไม่ทิ้งข้อมูลแม้แต่ตัวอักษรเดียว — เก็บ raw ดิบใน field `original_raw`
- field ที่ map ได้ → ขึ้น common + specs
- field ที่ map ไม่ได้ → เก็บใน `extra_fields` (ไม่ทิ้ง)
- เก็บ source_file, source_row, source_sheet เพื่อย้อนกลับดูได้
- ใช้ upsert — ถ้ามี document อยู่แล้ว (match ด้วย source_file + source_row) จะอัปเดต ไม่ทับซ้อน
- flag --reset สำหรับลบข้อมูลเดิมทั้งหมดก่อน import (ใช้เมื่อต้องการเริ่มใหม่)

Usage:
    .venv/bin/python scripts/import_adminbase.py
    .venv/bin/python scripts/import_adminbase.py --reset
    .venv/bin/python scripts/import_adminbase.py --dry-run   # อ่าน + แปลง แต่ไม่เขียน DB
"""

from __future__ import annotations

import argparse
import os
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from dotenv import load_dotenv
from pymongo import MongoClient

# ---- config ----

ROOT = Path(__file__).resolve().parent.parent
load_dotenv(ROOT / ".env")

ADMIN_DB_NAME = os.environ.get("ADMIN_MONGO_DB", "chatbot_admin").strip()
KB_COLLECTION = os.environ.get("ADMIN_MONGO_COLLECTION_KB", "knowledge_base").strip()
ADMINBASE_DIR = ROOT / "adminbase"

# ---- field mapping ----
# ชื่อ column ใน Excel → field ใน schema knowledge_base
# รองรับหลายชื่อ (เพราะไฟล์ตั้งชื่อไม่เหมือนกัน)

COMMON_FIELD_MAP: dict[str, list[str]] = {
    "brand": ["ชื่อแบรนด์", "แบรนด์", "brand"],
    "model": ["รุ่นสินค้า", "รุ่น", "model"],
    "category": ["ประเภท", "ประเภทสินค้า", "category"],
    "highlights": ["จุดเด่นสินค้า", "จุดเด่น"],
    "description": ["ข้อมูลสินค้า", "ข้อมูลสำคัญ", "ข้อมูลสินค้าหลังกล่อง", "ข้อมูล"],
    "box_contents": [
        "อุปกรณ์ที่ได้รับในแพ็กเกจ",
        "อุปกรณ์ภายในกล่อง",
        "อุปกรณ์ในกล่อง",
        "อุปกรณ์ที่แนะนำ",
    ],
    "warranty_period": ["ระยะเวลาการรับประกัน", "ระยะเวลารับประกัน"],
    "warranty_note": ["การรับประกันสินค้า", "การรับประกัน"],
    "notes": ["หมายเหตุสำคัญ", "หมายเหตุ"],
    "weight": ["น้ำหนักสินค้า", "น้ำหนัก"],
    "dimensions": ["ขนาดสินค้า", "ขนาด"],
    "question": ["คำถามเกี่ยวสินค้า", "คำถาม"],
    "answer": ["คำตอบ", "คำตอบเกี่ยวกับสินค้า"],
}

# field ที่เป็น "สเปก" — เก็บใน specs (dynamic)
# ไม่ต้อง map ทุก field — ที่เหลือจะไป extra_fields อัตโนมัติ
KNOWN_COMMON_KEYS: set[str] = set()
for _vals in COMMON_FIELD_MAP.values():
    KNOWN_COMMON_KEYS.update(v.lower() for v in _vals)

# field ที่ไม่ใช่สเปก แต่ก็ไม่ใช่ common — ไป extra_fields
# (เช่น วิธีการใช้งาน, การดูแลรักษา, วิธีการใส่สายนาฬิกา)
# ไม่ต้องระบุ — ทุก field ที่ไม่ใช่ common จะไป extra_fields อัตโนมัติ


def _build_reverse_map() -> dict[str, str]:
    """สร้าง reverse map: 'ชื่อ column ใน excel (lower)' → 'field ใน schema'."""
    rev: dict[str, str] = {}
    for schema_field, excel_names in COMMON_FIELD_MAP.items():
        for name in excel_names:
            rev[name.strip().lower()] = schema_field
    return rev


REVERSE_MAP = _build_reverse_map()


def _cell_to_str(val) -> str:
    """แปลงค่าจาก Excel เป็น string โดยไม่ทิ้งข้อมูล."""
    if val is None:
        return ""
    if isinstance(val, str):
        return val
    if isinstance(val, float):
        # ถ้าเป็นจำนวนเต็ม ไม่ต้องมี .0
        if val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, bool):
        return "true" if val else "false"
    if isinstance(val, datetime):
        return val.isoformat()
    return str(val)


def _detect_category_id(category: str, brand: str, model: str) -> str:
    """แปลง category ภาษาไทย/อังกฤษ → slug สำหรับ category_id."""
    low = (category or "").lower()
    combined = f"{category} {model}".lower()
    # ลำดับสำคัญ — เช็คเฉพาะก่อน
    if any(k in low for k in ["หูฟัง", "earphone", "earbuds", "tws", "headphone"]):
        return "earphone"
    if any(k in low for k in ["พาวเวอร์แบงค์", "powerbank", "power bank", "แบตสำรอง"]):
        return "powerbank"
    if any(k in low for k in ["สมาร์ทวอช", "smartwatch", "smart watch", "นาฬิกา"]):
        return "smartwatch"
    if any(k in low for k in ["โทรศัพท์", "มือถือ", "phone", "smartphone", "โทสับ"]):
        return "phone"
    if any(k in low for k in ["พัดลม", "fan", "cooling"]):
        return "fan"
    if any(k in low for k in ["ชาร์จ", "charger", "หัวชาร์จ", "สายชาร์จ", "cable"]):
        return "charger"
    if any(k in low for k in ["กล้องวงจรปิด", "cctv", "camera", "กล้องติดรถ", "dash cam"]):
        return "camera"
    if any(k in low for k in ["ลำโพง", "speaker", "bluetooth speak"]):
        return "speaker"
    if any(k in low for k in ["เครื่องนวด", "massager", "massager"]):
        return "massager"
    if any(k in low for k in ["เครื่องดูดฝุ่น", "vacuum"]):
        return "vacuum"
    if any(k in low for k in ["เครื่องฟอกอากาศ", "air purifier"]):
        return "air_purifier"
    if any(k in low for k in ["เครื่องทำความชื้น", "humidifier"]):
        return "humidifier"
    if any(k in low for k in ["ทำเล็บ", "nail"]):
        return "nail_care"
    if any(k in low for k in ["โกนหนวด", "shaver", "shave"]):
        return "shaver"
    if any(k in low for k in ["โต๊ะ", "table", "desk"]):
        return "table"
    if any(k in low for k in ["โซฟา", "sofa"]):
        return "sofa"
    if any(k in low for k in ["หมอน", "pillow"]):
        return "pillow"
    if any(k in low for k in ["เบาะ", "cushion"]):
        return "cushion"
    if any(k in low for k in ["เครื่องชั่ง", "scale"]):
        return "scale"
    if any(k in low for k in ["ssd", "drive", "ไดร์"]):
        return "storage"
    if any(k in low for k in ["แท็บเล็ต", "tablet", "pad"]):
        return "tablet"
    if any(k in low for k in ["ไมค์", "ไมโครโฟน", "microphone"]):
        return "microphone"
    if any(k in low for k in ["แท็ก", "tag", "tracker"]):
        return "tracker"
    if any(k in low for k in ["ถ่าน", "battery", "แบตเตอรี่"]):
        return "battery"
    if any(k in low for k in ["เครื่องใช้ไฟฟ้า", "appliance"]):
        return "appliance"
    if any(k in low for k in ["เครื่องใช้ภายในบ้าน", "home"]):
        return "home_appliance"
    if any(k in low for k in ["ลู่วิ่ง", "treadmill"]):
        return "treadmill"
    if any(k in low for k in ["ไดร์เป่าผม", "hair dryer"]):
        return "hair_dryer"
    if any(k in low for k in ["สายรัด", "strap", "nylon"]):
        return "strap"
    if any(k in low for k in ["เครื่องทำเล็บ", "nail"]):
        return "nail_care"
    if any(k in low for k in ["พยุงพุง", "พยุง"]):
        return "massager"
    if any(k in low for k in ["เครื่องโกน", "โกน"]):
        return "shaver"
    # fallback
    return "other"


def _is_qa_file(filename: str) -> bool:
    """ไฟล์ที่ชื่อบอกว่าเป็น Q&A."""
    low = filename.lower()
    return "ถาม" in low and "ตอบ" in low


def _is_comparison_file(filename: str) -> bool:
    """ไฟล์ที่เป็นตารางเปรียบเทียบ."""
    low = filename.lower()
    return "เปรียบเทียบ" in low or "หัวข้อเปรียบเทียบ" in low


def _is_spec_file(filename: str, sheet_name: str) -> bool:
    """ไฟล์ที่เป็น spec sheet (สเปกละเอียด)."""
    low = (filename + " " + sheet_name).lower()
    return "spec" in low


def _parse_excel_row(
    header: list[str],
    row: tuple,
    source_file: str,
    source_row: int,
    source_sheet: str,
) -> dict | None:
    """แปลง row จาก Excel → document ตาม schema knowledge_base.

    คืน None ถ้า row ว่างทั้งหมด.
    """
    # สร้าง raw dict (เก็บทุก column ตามชื่อเดิม)
    raw: dict[str, str] = {}
    for i, col_name in enumerate(header):
        if not col_name:
            continue
        val = row[i] if i < len(row) else None
        cell_str = _cell_to_str(val)
        if cell_str:
            raw[col_name] = cell_str

    if not raw:
        return None  # row ว่าง

    # แปลง raw → common fields
    doc: dict = {
        "type": "product_spec",
        "brand": "",
        "model": "",
        "category": "",
        "category_id": "",
        "highlights": "",
        "description": "",
        "box_contents": "",
        "warranty_period": "",
        "warranty_note": "",
        "notes": "",
        "weight": "",
        "dimensions": "",
        "specs": {},
        "extra_fields": {},
        "original_raw": raw,
        "source_file": source_file,
        "source_row": source_row,
        "source_sheet": source_sheet,
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
        "updated_by": "system_import",
        "version": 1,
        "active": True,
    }

    # map common fields
    used_raw_keys: set[str] = set()
    for raw_key, raw_val in raw.items():
        schema_field = REVERSE_MAP.get(raw_key.strip().lower())
        if schema_field and schema_field in doc:
            # field ที่เป็น string ใน doc
            if isinstance(doc[schema_field], str):
                doc[schema_field] = raw_val
            used_raw_keys.add(raw_key)

    # field ที่เหลือ → specs หรือ extra_fields
    # ถ้าเป็น field ที่ดูเหมือนสเปก (มีค่าสั้นๆ) → specs
    # ถ้าเป็น field ที่ดูเหมือนคำอธิบาย (ยาว) → extra_fields
    for raw_key, raw_val in raw.items():
        if raw_key in used_raw_keys:
            continue
        # ถ้าค่ายาวเกิน 200 ตัวอักษร → extra_fields (เป็นคำอธิบาย)
        # ถ้าสั้น → specs
        if len(raw_val) > 200:
            doc["extra_fields"][raw_key] = raw_val
        else:
            doc["specs"][raw_key] = raw_val

    # detect category_id
    doc["category_id"] = _detect_category_id(
        doc.get("category", ""), doc.get("brand", ""), doc.get("model", "")
    )

    # ถ้าเป็นไฟล์ Q&A → เปลี่ยน type
    if _is_qa_file(source_file):
        doc["type"] = "qa"
        if doc.get("question") or doc.get("answer"):
            doc["question"] = doc.get("question", "")
            doc["answer"] = doc.get("answer", "")

    # ถ้าเป็นไฟล์เปรียบเทียบ → เปลี่ยน type
    if _is_comparison_file(source_file):
        doc["type"] = "comparison"

    # ถ้าเป็น spec file → type = product_spec (default อยู่แล้ว)
    # แต่เก็บ flag ว่าเป็น spec
    if _is_spec_file(source_file, source_sheet):
        doc["is_spec_sheet"] = True

    return doc


def _parse_txt_file(filepath: Path) -> list[dict]:
    """อ่านไฟล์ .txt (เงื่อนไขรับประกันทั่วไป) → document type=general_faq."""
    content = filepath.read_text(encoding="utf-8", errors="replace").strip()
    if not content:
        return []

    doc = {
        "type": "general_faq",
        "topic": "รับประกัน",
        "question_patterns": ["รับประกัน", "เคลม", "warranty", "garantee", "guarantee"],
        "answer": content,
        "applies_to_brands": [],
        "applies_to_categories": [],
        "source_file": filepath.name,
        "source_row": 1,
        "source_sheet": "",
        "original_raw": {"content": content},
        "created_at": datetime.now(timezone.utc),
        "updated_at": datetime.now(timezone.utc),
        "updated_by": "system_import",
        "version": 1,
        "active": True,
    }
    return [doc]


def _get_admin_db() -> "MongoClient":
    """เชื่อม DB admin โดยใช้ค่าจาก .env."""
    uri = os.environ.get("ADMIN_MONGO_URI", "").strip()
    if uri:
        client = MongoClient(uri)
    else:
        host = os.environ.get("ADMIN_MONGO_HOST", "127.0.0.1:27017").strip()
        username = os.environ.get("ADMIN_MONGO_USERNAME", "").strip()
        password = os.environ.get("ADMIN_MONGO_PASSWORD", "").strip()
        auth_source = os.environ.get("ADMIN_MONGO_AUTH_SOURCE", "admin").strip()
        tls = os.environ.get("ADMIN_MONGO_TLS", "false").strip().lower() == "true"

        params: dict = {"host": host, "authSource": auth_source, "tls": tls}
        if username:
            params["username"] = username
        if password:
            params["password"] = password
        client = MongoClient(**params)

    return client


def main() -> int:
    parser = argparse.ArgumentParser(description="Import adminbase Excel → MongoDB")
    parser.add_argument(
        "--reset",
        action="store_true",
        help="ลบข้อมูลเดิมใน knowledge_base ทั้งหมดก่อน import",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="อ่าน + แปลง แต่ไม่เขียน DB (สำหรับทดสอบ)",
    )
    parser.add_argument(
        "--file",
        type=str,
        default=None,
        help="import เฉพาะไฟล์ที่ระบุ (ชื่อไฟล์ใน adminbase/)",
    )
    args = parser.parse_args()

    if not ADMINBASE_DIR.exists():
        print(f"ERROR: adminbase dir not found: {ADMINBASE_DIR}")
        return 1

    # หาไฟล์ทั้งหมด
    xlsx_files = sorted(ADMINBASE_DIR.glob("*.xlsx"))
    txt_files = sorted(ADMINBASE_DIR.glob("*.txt"))

    if args.file:
        xlsx_files = [f for f in xlsx_files if f.name == args.file]
        txt_files = [f for f in txt_files if f.name == args.file]
        if not xlsx_files and not txt_files:
            print(f"ERROR: file not found: {args.file}")
            return 1

    print(f"=== Import adminbase → MongoDB ===")
    print(f"  DB: {ADMIN_DB_NAME}")
    print(f"  Collection: {KB_COLLECTION}")
    print(f"  Source: {ADMINBASE_DIR}")
    print(f"  Excel files: {len(xlsx_files)}")
    print(f"  TXT files: {len(txt_files)}")
    print(f"  Mode: {'dry-run' if args.dry_run else 'write'}")
    if args.reset:
        print(f"  Reset: YES (will delete existing data)")
    print()

    # อ่าน + แปลงทุกไฟล์
    all_docs: list[dict] = []
    errors: list[str] = []
    file_stats: list[tuple[str, int, int]] = []  # (filename, rows_read, docs_produced)

    for f in xlsx_files:
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            file_docs = 0
            file_rows = 0
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                header = [
                    str(c).strip() if c is not None else "" for c in rows[0]
                ]
                for row_idx, row in enumerate(rows[1:], start=2):
                    file_rows += 1
                    doc = _parse_excel_row(
                        header=header,
                        row=row,
                        source_file=f.name,
                        source_row=row_idx,
                        source_sheet=sheet_name,
                    )
                    if doc:
                        all_docs.append(doc)
                        file_docs += 1
            wb.close()
            file_stats.append((f.name, file_rows, file_docs))
        except Exception as exc:
            errors.append(f"{f.name}: {exc}")
            file_stats.append((f.name, 0, 0))

    for f in txt_files:
        try:
            docs = _parse_txt_file(f)
            all_docs.extend(docs)
            file_stats.append((f.name, 0, len(docs)))
        except Exception as exc:
            errors.append(f"{f.name}: {exc}")
            file_stats.append((f.name, 0, 0))

    # สรุปการอ่าน
    print("=== File stats ===")
    print(f"{'File':<50s} {'rows':>5s} {'docs':>5s}")
    print("-" * 65)
    total_rows = 0
    for fname, rows, docs in file_stats:
        print(f"{fname[:48]:<50s} {rows:5d} {docs:5d}")
        total_rows += rows
    print("-" * 65)
    print(f"{'TOTAL':<50s} {total_rows:5d} {len(all_docs):5d}")
    print()

    if errors:
        print("=== Errors ===")
        for e in errors:
            print(f"  {e}")
        print()

    # สรุป type distribution
    type_counts: dict[str, int] = {}
    for d in all_docs:
        t = d.get("type", "unknown")
        type_counts[t] = type_counts.get(t, 0) + 1
    print("=== Type distribution ===")
    for t, c in sorted(type_counts.items()):
        print(f"  {t}: {c}")
    print()

    # สรุป category distribution
    cat_counts: dict[str, int] = {}
    for d in all_docs:
        if d.get("type") == "product_spec":
            cat = d.get("category_id", "other")
            cat_counts[cat] = cat_counts.get(cat, 0) + 1
    print("=== Category distribution (product_spec) ===")
    for cat, c in sorted(cat_counts.items(), key=lambda x: -x[1]):
        print(f"  {cat}: {c}")
    print()

    # ตรวจว่าไม่หล่นข้อมูล — ทุก doc ต้องมี original_raw ที่ไม่ว่าง
    empty_raw = sum(1 for d in all_docs if not d.get("original_raw"))
    print(f"=== Data integrity ===")
    print(f"  Documents with empty original_raw: {empty_raw} (should be 0)")
    print()

    if args.dry_run:
        print("=== Dry run — not writing to DB ===")
        # แสดงตัวอย่าง 1 doc
        if all_docs:
            import json

            sample = all_docs[0].copy()
            # ตัดทอนให้ดูง่าย
            for k in ["original_raw", "extra_fields", "specs"]:
                if sample.get(k):
                    s = str(sample[k])
                    sample[k] = s[:200] + "..." if len(s) > 200 else s
            sample["created_at"] = str(sample["created_at"])
            sample["updated_at"] = str(sample["updated_at"])
            print(f"  Sample doc (first):")
            print(json.dumps(sample, ensure_ascii=False, indent=2))
        return 0

    # เขียนลง DB
    print("=== Writing to MongoDB ===")
    client = _get_admin_db()
    db = client[ADMIN_DB_NAME]
    coll = db[KB_COLLECTION]

    if args.reset:
        deleted = coll.delete_many({})
        print(f"  Reset: deleted {deleted.deleted_count} existing documents")

    # upsert — match ด้วย source_file + source_row
    inserted = 0
    updated = 0
    for doc in all_docs:
        filter_q = {
            "source_file": doc["source_file"],
            "source_row": doc["source_row"],
        }
        result = coll.replace_one(filter_q, doc, upsert=True)
        if result.upserted_id:
            inserted += 1
        else:
            updated += 1

    print(f"  Inserted: {inserted}")
    print(f"  Updated:  {updated}")
    print(f"  Total in collection now: {coll.count_documents({})}")
    client.close()

    print()
    print("=== Done ===")
    return 0


if __name__ == "__main__":
    sys.exit(main())
